import openpyxl
import json
import os

wb = openpyxl.load_workbook('CS.xlsx', data_only=True)
db_data = {
    'productos': [],
    'clientes': [],
    'proveedores': [],
    'inventario': [],
    'compras': [],
    'productosComprados': [],
    'ventas': [],
    'defectuosos': []
}

def to_str(v):
    if v is None: return None
    if hasattr(v, 'isoformat'): return v.isoformat()
    return v

def clean_val(v):
    return to_str(v)

try:
    for r in wb['Productos'].iter_rows(min_row=2, max_col=1, values_only=True):
        if r[0]: db_data['productos'].append({'nombre': clean_val(r[0])})
except Exception as e: print("Error Productos:", e)

try:
    for r in wb['Clientes'].iter_rows(min_row=2, max_col=3, values_only=True):
        if r[0]: db_data['clientes'].append({'nombre': clean_val(r[0]), 'celular': clean_val(r[1]), 'debe': clean_val(r[2])})
except Exception as e: print("Error Clientes:", e)

try:
    for r in wb['Proveedores'].iter_rows(min_row=2, max_col=4, values_only=True):
        if r[0]: db_data['proveedores'].append({'nombre': clean_val(r[0]), 'celular': clean_val(r[1]), 'url': clean_val(r[2]), 'etiquetas': clean_val(r[3])})
except Exception as e: print("Error Proveedores:", e)

try:
    for r in wb['Inventario'].iter_rows(min_row=2, max_col=4, values_only=True):
        if r[0]: db_data['inventario'].append({'nombre': clean_val(r[0]), 'cantidad': clean_val(r[1]), 'valorVenta': clean_val(r[2]), 'valorTotal': clean_val(r[3])})
except Exception as e: print("Error Inventario:", e)

try:
    for r in wb['Compras'].iter_rows(min_row=2, max_col=8, values_only=True):
        if r[0]: db_data['compras'].append({
            'numero': clean_val(r[0]), 'proveedor': clean_val(r[1]), 'valor': clean_val(r[2]),
            'fechaPedido': clean_val(r[3]), 'fechaLlegada': clean_val(r[4]),
            'descripcion': clean_val(r[5]), 'comentarios': clean_val(r[6]), 'factura': clean_val(r[7])
        })
except Exception as e: print("Error Compras:", e)

try:
    for r in wb['Productos Comprados'].iter_rows(min_row=2, max_col=5, values_only=True):
        if r[0] and r[1]: db_data['productosComprados'].append({'compraId': clean_val(r[0]), 'nombre': clean_val(r[1]), 'cantidad': clean_val(r[2]), 'valorUnitario': clean_val(r[3]), 'valorTotal': clean_val(r[4])})
except Exception as e: print("Error Productos Comprados:", e)

try:
    for r in wb['Ventas'].iter_rows(min_row=2, max_col=8, values_only=True):
        if r[0]: db_data['ventas'].append({
            'producto': clean_val(r[0]), 'cliente': clean_val(r[1]), 'cantidad': clean_val(r[2]),
            'valorPagado': clean_val(r[3]), 'valorVenta': clean_val(r[4]), 'valorTotal': clean_val(r[5]),
            'debe': clean_val(r[6]), 'comentarios': clean_val(r[7])
        })
except Exception as e: print("Error Ventas:", e)

if 'Productos Defectuosos' in wb.sheetnames:
    try:
        for r in wb['Productos Defectuosos'].iter_rows(min_row=2, max_col=5, values_only=True):
            if r[0]: db_data['defectuosos'].append({'compra': clean_val(r[0]), 'producto': clean_val(r[1]), 'cantidad': clean_val(r[2]), 'valorUnitario': clean_val(r[3]), 'valorTotal': clean_val(r[4])})
    except Exception as e: print("Error Defectuosos:", e)

os.makedirs('assets', exist_ok=True)
with open('assets/seed.json', 'w', encoding='utf-8') as f:
    json.dump(db_data, f, ensure_ascii=False, indent=2)

print("Created seed.json!")
